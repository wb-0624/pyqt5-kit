from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook


class ExcelKit:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.excel_name = excel_path.split('/')[-1]
        self.excel = None

    def read_excel(self):
        if self.excel_name.endswith('.xls'):
            self.read_xls()
        elif self.excel_name.endswith('.xlsx'):
            self.read_xlsx()
        else:
            raise Exception('File type error!')

    # todo
    def read_xls(self):
        raise Exception('xls is not supported!')

    def read_xlsx(self):
        self.excel = load_workbook(self.excel_path)

    def write_excel(self, data):
        if self.excel_name.endswith('.xls'):
            self.write_xls(data)
        elif self.excel_name.endswith('.xlsx'):
            self.write_xlsx(data)
        else:
            raise Exception('File type error!')

    # todo
    def write_xls(self, data):
        raise Exception('xls is not supported!')

    def write_xlsx(self, data):
        if self.excel is None:
            self.create_xlsx()
            self.excel = load_workbook(self.excel_path)
        ws = self.excel.active
        ws.append(data)

    def create_excel(self):
        if self.excel_name.endswith('.xls'):
            self.create_xls()
        elif self.excel_name.endswith('.xlsx'):
            self.create_xlsx()
        else:
            raise Exception('File type error!')

    def create_xls(self):
        raise Exception('xls is not supported!')

    def create_xlsx(self):
        self.excel = Workbook()
        self.excel.save(self.excel_path)

    def save(self):
        self.excel.save(self.excel_path)


if __name__ == '__main__':
    excel_kit = ExcelKit('题库.xlsx')
    excel_kit.write_excel([1, 2, 3, 4, 5])
    excel_kit.save()
